import os
import time
from datetime import datetime

import torchaudio

from tortoise.api import TextToSpeech
from tortoise.custom_models import Experimentation
from tortoise.utils.audio import load_voice
from openpyxl import Workbook, load_workbook


def write_txt_stat_file(duration_exp, duration_loop, exp, list_duration, output_path):
    with open(os.path.join(output_path, "exp_info.txt"), "w") as f:
        f.write(str(exp))
        f.write(f"\nexperimentation took: {duration_exp} \n")
        total_number_of_cars = sum(len(text) for text in exp.texts)
        f.write(
            f"\ntotal nb of cars: {total_number_of_cars}, car/s: {round(total_number_of_cars / duration_exp, 2)} \n\n")
        for i, result in enumerate(list_duration, start=1):
            f.write(
                f"text {i} took {str(result)} with {len(exp.texts[i - 1])}, car/s: {round(len(exp.texts[i - 1]) / duration_loop, 2)} \n")


def write_excel_stat_file(voice_name, texts, result_dir, tts_init, preset, all_params, duration_exp, duration_loop, exp, list_duration, output_path):
    output_file_name = os.path.join("results", "results.xlsx")

    if os.path.isfile(output_file_name):
        workbook = load_workbook(filename=output_file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active

    row = sheet.max_row + 1

    sheet[f'A{row}'] = 'fork'
    sheet[f'B{row}'] = "\n".join(texts)
    sheet[f'C{row}'] = voice_name
    sheet[f'D{row}'] = result_dir

    sheet[f'F{row}'] = tts_init.get("use_deepspeed", "")
    sheet[f'G{row}'] = tts_init.get("kv_cache", "")
    sheet[f'H{row}'] = tts_init.get("half", "")
    sheet[f'I{row}'] = "v3"
    sheet[f'J{row}'] = preset
    sheet[f'K{row}'] = all_params.get("temperature", "")

    for i, duration in enumerate(list_duration, start=13):
        sheet.cell(row=row, column=i, value=duration)

    total_number_of_cars = sum(len(text) for text in texts)

    sheet.cell(row=row, column=14 + len(list_duration) , value=round(total_number_of_cars / duration_exp, 2))
    workbook.save(filename=output_file_name)

def generate_sentence_and_save(experimentations: Experimentation):
    voice_name = ""
    for i, exp in enumerate(experimentations, start=1):
        start_time_exp = time.time()
        total_experimentations = len(experimentations)
        print(f"exp {i} on {total_experimentations}")

        now = datetime.now()
        formatted_now = now.strftime('%Y%m%d_%H_%M_%S') + '_' + str(int(now.microsecond / 1000)).zfill(3)

        list_duration = []
        if not exp.voice_name == voice_name:
            voice_samples, conditioning_latents = load_voice(exp.voice_name)
            voice_name = exp.voice_name

        result_dir = f"{formatted_now}_{exp.voice_name}"
        output_path = os.path.join("results", result_dir)
        os.makedirs(output_path, exist_ok=True)

        tts = TextToSpeech(**exp.tts_init)
        print("tts initialized")

        for i, text in enumerate(exp.texts, start=1):
            start_time_loop = time.time()
            print(f"sentence number {i} in progress on {len(exp.texts)}")

            all_params = {**exp.parameters_preset, **exp.parameters}
            print(all_params)

            gen = tts.tts(text, voice_samples=voice_samples, conditioning_latents=conditioning_latents, **all_params)

            torchaudio.save(os.path.join(output_path, f'{exp.voice_name}_{i}.wav'), gen.squeeze(0).cpu(), 24000)
            end_time_loop = time.time()
            duration_loop = round(end_time_loop - start_time_loop, 2)

            print(f"sentence number {i} took {duration_loop} with nb car {len(text)}")
            list_duration.append(duration_loop)

        end_time_exp = time.time()
        duration_exp = round(end_time_exp - start_time_exp, 2)
        print(f"experimentation took {duration_exp}")

        write_txt_stat_file(duration_exp, duration_loop, exp, list_duration, output_path)
        write_excel_stat_file(exp.voice_name, exp.texts,result_dir, exp.tts_init, exp.preset, all_params, duration_exp,
                            duration_loop, exp, list_duration, output_path)
